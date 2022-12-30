from pprint import pprint

from openpyxl import load_workbook,Workbook


def load_case_by_excel(file):
    """
    从excel表中加载用例,这时读取到的还是数据，不是真正的用例并可读取excel中的sheets的title
    生成真正的用例要通过suite_list，传入create_case_by_excel（suite_list）生成真正的用例Test_list
    :param file:  文件
    :return:  suite_list  #最终结构==>suite_list = ["case_list":[{用例case}],'name': 'Sheet1',["case_list":[{用例case}],'name': 'Sheet2']
    """
    wb: Workbook = load_workbook(file)   #openpyxl内置的打开文件方法
    # ws = wb.active  #使用默认的工作表
    #读取到的用例信息生成用例保存下来
    suite_list = []   #最终结构==>suite_list = ["case_list":[{用例case}],'name': 'Sheet1',["case_list":[{用例case}],'name': 'Sheet2']

    for ws in wb.worksheets:    #读取xlsl表中的所有工作页worksheets
        case_list = []  #存放用例的列表
        suite = {
            "name": ws.title,  #工作表的名字
            "case_list": case_list,  #放入字典
        }
        suite_list.append(suite)
        for line in ws.iter_rows(values_only=True, min_row=2):  #从工作表中逐行读取内容;values_only=True为只要excel里面的值,min_row=2:从最小行开始
            if line[0] == 0:   #若第一列值为0，说明进入新的用例
                case = {   #创建新用例
                    "name": "",  #用例名
                    "steps": []  #用例步骤
                }
                case_list.append(case)
                case["name"] = line[3]
            else:
                case["steps"].append(line)
    return(suite_list)

#调试用：
# cases = load_case_by_excel("../testcases/testcases.xlsx")
# print(cases)
# pprint(cases[0]['case_list'])